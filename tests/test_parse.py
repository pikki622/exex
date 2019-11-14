import unittest

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from exex import parse


class ParseTest(unittest.TestCase):
    book: Workbook = None
    sheet: Worksheet = None

    def setUp(self) -> None:
        self.book = load_workbook("tests/files/sample.xlsx")
        self.sheet = self.book.active

    def tearDown(self) -> None:
        del self.sheet
        self.book.close()
        del self.book

    def test_from_cell__name(self):
        self.assertEqual(parse.from_cell(self.sheet["A1"]), "name")

    def test_from_cell__coordinate(self):
        self.assertEqual(parse.from_cell(self.sheet.cell(1, 1)), "name")

    def test_from_array__row(self):
        self.assertEqual(
            parse.from_array(self.sheet[1]), ["name", "abbreviation", "age"]
        )

    def test_from_array__column(self):
        self.assertEqual(
            parse.from_array(self.sheet["A"]), ["name", "alpha", "beta", "gamma"]
        )

    def test_from_range__values(self):
        self.assertEqual(
            parse.from_range(self.sheet.values),
            [
                ["name", "abbreviation", "age"],
                ["alpha", "a", 1],
                ["beta", "b", 2],
                ["gamma", "g", 3],
            ],
        )

    def test_from_range__horizontal_cell_range(self):
        self.assertEqual(
            parse.from_range(self.sheet["A1:B1"]), [["name", "abbreviation"]]
        )

    def test_from_range__vertical_cell_Range(self):
        self.assertEqual(
            parse.from_range(self.sheet["A1:A4"]),
            [["name"], ["alpha"], ["beta"], ["gamma"]],
        )

    def test_from_range__rectangular_cell_range(self):
        self.assertEqual(
            parse.from_range(self.sheet["A1:B2"]),
            [["name", "abbreviation"], ["alpha", "a"]],
        )

    def test_from_range__row_range(self):
        self.assertEqual(
            parse.from_range(self.sheet[2:3]), [["alpha", "a", 1], ["beta", "b", 2]]
        )

    def test_from_range__column_range(self):
        self.assertEqual(
            parse.from_range(self.sheet["A:B"]),
            [["name", "alpha", "beta", "gamma"], ["abbreviation", "a", "b", "g"]],
        )

    def test_parse_values(self):
        self.assertEqual(parse.values(None), None)
        self.assertEqual(parse.values(False), False)
        self.assertEqual(parse.values(5), 5)
        self.assertEqual(parse.values("a"), "a")
        self.assertEqual(parse.values([1, 2, 3]), [1, 2, 3])

        self.assertEqual(parse.values(self.sheet["A1"]), "name")
        self.assertEqual(parse.values(self.sheet.cell(1, 1)), "name")
        self.assertEqual(parse.values(self.sheet[1]), ["name", "abbreviation", "age"])
        self.assertEqual(
            parse.values(self.sheet["A"]), ["name", "alpha", "beta", "gamma"]
        )
        self.assertEqual(
            parse.values(self.sheet.values),
            [
                ["name", "abbreviation", "age"],
                ["alpha", "a", 1],
                ["beta", "b", 2],
                ["gamma", "g", 3],
            ],
        )
        self.assertEqual(parse.values(self.sheet["A1:B1"]), [["name", "abbreviation"]])
        self.assertEqual(
            parse.values(self.sheet["A1:A4"]),
            [["name"], ["alpha"], ["beta"], ["gamma"]],
        )
        self.assertEqual(
            parse.values(self.sheet["A1:B2"]),
            [["name", "abbreviation"], ["alpha", "a"]],
        )
        self.assertEqual(
            parse.values(self.sheet[2:3]), [["alpha", "a", 1], ["beta", "b", 2]]
        )
        self.assertEqual(
            parse.values(self.sheet["A:B"]),
            [["name", "alpha", "beta", "gamma"], ["abbreviation", "a", "b", "g"]],
        )
